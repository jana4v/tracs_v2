from typing import Optional
from datetime import datetime, timezone
from pathlib import Path
import shutil
import subprocess
import tempfile

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from src.database.connection import (
    get_calibration_data_collection,
    get_calibration_runs_collection,
    get_configuration_collection,
    get_project_instruments_collection,
    get_project_power_meters_collection,
    get_project_tsm_paths_collection,
    get_transmitter_misc_collection,
    get_transmitters_collection,
    get_instruments_collection,
)
from src.config import settings
from src.calibration.base import CalibrationDependencies
from src.database.connection import Database
from src.database.sqlite_json_store import SQLiteJsonCollection
from src.repositories.cal_sg_calibration_repo import CalSgCalibrationRepository
from src.repositories.downlink_cal_calibration_repo import DownlinkCalCalibrationRepository
from src.repositories.inject_cal_calibration_repo import InjectCalCalibrationRepository
from src.repositories.calibration_data_repo import CalibrationDataRepository
from src.repositories.env_data_repo import EnvDataRepository
from src.repositories.test_systems_repo import TestSystemsRepository
from src.repositories.test_phases_repo import TestPhasesRepository
from src.repositories.transmitter_repo import TransmitterRepository
from src.schemas.calibration_data import (
    CalSgCompletedFrequenciesResponse,
    CalSgDataRowsResponse,
    DownlinkCalDataRowsResponse,
    CalIdsResponse,
    CalibrationReportGenerateRequest,
    CalibrationReportGenerateResponse,
    CalibrationRunAbortResponse,
    CalibrationRunPromptResponseRequest,
    CalibrationRunSnapshot,
    CalibrationRunStartRequest,
    MeasureRunStartRequest,
    MeasureRunStartResponse,
    MeasureOptionsResponse,
)
from src.services.calibration_run_service import CalibrationRunService
from src.measurements import MeasureRunService

router = APIRouter(prefix="/api/v2", tags=["Calibration Data"])
_run_service: Optional[CalibrationRunService] = None
_measure_service: Optional[MeasureRunService] = None


def get_repo(
    collection: SQLiteJsonCollection = Depends(get_calibration_data_collection),
) -> CalibrationDataRepository:
    return CalibrationDataRepository(collection)


def get_calibration_dependencies(
    runs_collection: SQLiteJsonCollection = Depends(get_calibration_runs_collection),
    transmitters_collection: SQLiteJsonCollection = Depends(get_transmitters_collection),
    tsm_paths_collection: SQLiteJsonCollection = Depends(get_project_tsm_paths_collection),
    misc_collection: SQLiteJsonCollection = Depends(get_transmitter_misc_collection),
    instruments_collection: SQLiteJsonCollection = Depends(get_instruments_collection),
    project_instruments_collection: SQLiteJsonCollection = Depends(get_project_instruments_collection),
    project_power_meters_collection: SQLiteJsonCollection = Depends(get_project_power_meters_collection),
    configuration_collection: SQLiteJsonCollection = Depends(get_configuration_collection),
) -> CalibrationDependencies:
    _ = runs_collection
    return CalibrationDependencies(
        transmitter_repo=TransmitterRepository(transmitters_collection, tsm_paths_collection, misc_collection),
        test_systems_repo=TestSystemsRepository(
            transmitters_collection=transmitters_collection,
            instruments_collection=instruments_collection,
            project_instruments_collection=project_instruments_collection,
            project_power_meters_collection=project_power_meters_collection,
            project_tsm_paths_collection=tsm_paths_collection,
            configuration_collection=configuration_collection,
        ),
        cal_sg_repo=CalSgCalibrationRepository(Database._db_path, settings.CAL_SG_CALIBRATION_TABLE),
        inject_cal_repo=InjectCalCalibrationRepository(Database._db_path, settings.INJECT_CAL_CALIBRATION_TABLE),
        downlink_cal_repo=DownlinkCalCalibrationRepository(Database._db_path, settings.DOWNLINK_CAL_CALIBRATION_TABLE),
    )


def get_run_service(
    runs_collection: SQLiteJsonCollection = Depends(get_calibration_runs_collection),
    dependencies: CalibrationDependencies = Depends(get_calibration_dependencies),
) -> CalibrationRunService:
    global _run_service
    if _run_service is None:
        _run_service = CalibrationRunService(runs_collection, dependencies)
    return _run_service


def get_measure_service() -> MeasureRunService:
    global _measure_service
    if _measure_service is None:
        _measure_service = MeasureRunService()
    return _measure_service


@router.get("/calibration/cal-ids", response_model=CalIdsResponse)
def get_cal_ids(
    cal_type: Optional[str] = None,
    repo: CalibrationDataRepository = Depends(get_repo),
):
    """Return distinct cal_ids from CalibrationData collection.
    Optionally filter by cal_type (uplink | downlink | fixedpad | calSG | injectcal).
    """
    ids = repo.get_cal_ids(cal_type=cal_type)
    return CalIdsResponse(cal_ids=ids)


@router.get("/measure/options", response_model=MeasureOptionsResponse)
def get_measure_options(
    repo: CalibrationDataRepository = Depends(get_repo),
):
    test_phases_repo = TestPhasesRepository(Database._db_path, settings.TEST_PHASES_TABLE)
    downlink_repo = DownlinkCalCalibrationRepository(Database._db_path, settings.DOWNLINK_CAL_CALIBRATION_TABLE)

    test_phases_rows = test_phases_repo.list_rows()
    test_phases = [str(r.get("TEST_PHASE") or "").strip() for r in test_phases_rows if str(r.get("TEST_PHASE") or "").strip() != ""]

    downlink_ids = downlink_repo.list_cal_ids()
    uplink_ids = repo.get_cal_ids(cal_type="uplink")
    legacy_downlink_ids = repo.get_cal_ids(cal_type="downlink")

    seen: set[str] = set()
    cal_ids: list[str] = []
    for cid in downlink_ids + uplink_ids + legacy_downlink_ids:
        normalized = str(cid or "").strip()
        if normalized == "" or normalized in seen:
            continue
        seen.add(normalized)
        cal_ids.append(normalized)

    default_cal_id = cal_ids[0] if len(cal_ids) > 0 else None
    return MeasureOptionsResponse(
        test_phases=test_phases,
        cal_ids=cal_ids,
        default_cal_id=default_cal_id,
    )


@router.post("/measure/runs/start", response_model=MeasureRunStartResponse)
async def start_measure_run(
    payload: MeasureRunStartRequest,
    service: MeasureRunService = Depends(get_measure_service),
    dependencies: CalibrationDependencies = Depends(get_calibration_dependencies),
):
    try:
        return await service.start_run(payload, dependencies)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/calibration/cal-sg/completed-frequencies", response_model=CalSgCompletedFrequenciesResponse)
def get_cal_sg_completed_frequencies(cal_id: str):
    cal_repo = CalSgCalibrationRepository(Database._db_path, settings.CAL_SG_CALIBRATION_TABLE)
    normalized_cal_id = str(cal_id or "").strip()
    if normalized_cal_id == "":
        raise HTTPException(status_code=400, detail="cal_id is required")
    frequencies = cal_repo.list_frequencies(normalized_cal_id)
    return CalSgCompletedFrequenciesResponse(cal_id=normalized_cal_id, frequencies=frequencies)


@router.get("/calibration/cal-sg/data", response_model=CalSgDataRowsResponse)
def get_cal_sg_data(cal_id: str):
    cal_repo = CalSgCalibrationRepository(Database._db_path, settings.CAL_SG_CALIBRATION_TABLE)
    normalized_cal_id = str(cal_id or "").strip()
    if normalized_cal_id == "":
        raise HTTPException(status_code=400, detail="cal_id is required")
    rows = cal_repo.list_rows(cal_id=normalized_cal_id)
    return CalSgDataRowsResponse(cal_id=normalized_cal_id, rows=rows)


@router.get("/calibration/inject-cal/completed-frequencies", response_model=CalSgCompletedFrequenciesResponse)
def get_inject_cal_completed_frequencies(cal_id: str):
    cal_repo = InjectCalCalibrationRepository(Database._db_path, settings.INJECT_CAL_CALIBRATION_TABLE)
    normalized_cal_id = str(cal_id or "").strip()
    if normalized_cal_id == "":
        raise HTTPException(status_code=400, detail="cal_id is required")
    frequencies = cal_repo.list_frequencies(normalized_cal_id)
    return CalSgCompletedFrequenciesResponse(cal_id=normalized_cal_id, frequencies=frequencies)


@router.get("/calibration/inject-cal/data", response_model=CalSgDataRowsResponse)
def get_inject_cal_data(cal_id: str):
    cal_repo = InjectCalCalibrationRepository(Database._db_path, settings.INJECT_CAL_CALIBRATION_TABLE)
    normalized_cal_id = str(cal_id or "").strip()
    if normalized_cal_id == "":
        raise HTTPException(status_code=400, detail="cal_id is required")
    rows = cal_repo.list_rows(cal_id=normalized_cal_id)
    return CalSgDataRowsResponse(cal_id=normalized_cal_id, rows=rows)


@router.get("/calibration/downlink-cal/data", response_model=DownlinkCalDataRowsResponse)
def get_downlink_cal_data(cal_id: str):
    cal_repo = DownlinkCalCalibrationRepository(Database._db_path, settings.DOWNLINK_CAL_CALIBRATION_TABLE)
    normalized_cal_id = str(cal_id or "").strip()
    if normalized_cal_id == "":
        raise HTTPException(status_code=400, detail="cal_id is required")
    rows = cal_repo.list_rows(cal_id=normalized_cal_id)
    return DownlinkCalDataRowsResponse(cal_id=normalized_cal_id, rows=rows)


@router.post("/calibration/runs/start", response_model=CalibrationRunSnapshot)
async def start_calibration_run(
    payload: CalibrationRunStartRequest,
    service: CalibrationRunService = Depends(get_run_service),
):
    try:
        return await service.start_run(payload)
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc


@router.get("/calibration/runs/active", response_model=Optional[CalibrationRunSnapshot])
async def get_active_calibration_run(
    service: CalibrationRunService = Depends(get_run_service),
):
    return await service.get_active_snapshot()


@router.get("/calibration/runs/latest", response_model=Optional[CalibrationRunSnapshot])
async def get_latest_calibration_run(
    cal_type: Optional[str] = None,
    service: CalibrationRunService = Depends(get_run_service),
):
    return await service.get_latest_snapshot(cal_type=cal_type)


@router.get("/calibration/runs/{run_id}", response_model=CalibrationRunSnapshot)
async def get_calibration_run(
    run_id: str,
    service: CalibrationRunService = Depends(get_run_service),
):
    snapshot = await service.get_snapshot(run_id)
    if snapshot is None:
        raise HTTPException(status_code=404, detail="Run not found")
    return snapshot


@router.post("/calibration/runs/{run_id}/prompt-response", response_model=CalibrationRunSnapshot)
async def post_operator_prompt_response(
    run_id: str,
    payload: CalibrationRunPromptResponseRequest,
    service: CalibrationRunService = Depends(get_run_service),
):
    try:
        return await service.operator_prompt_response(run_id, payload)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc


@router.post("/calibration/runs/{run_id}/abort", response_model=CalibrationRunAbortResponse)
async def abort_calibration_run(
    run_id: str,
    service: CalibrationRunService = Depends(get_run_service),
):
    return await service.abort_run(run_id)


@router.get("/calibration/runs/{run_id}/events")
async def stream_calibration_run_events(
    run_id: str,
    service: CalibrationRunService = Depends(get_run_service),
):
    return StreamingResponse(
        service.stream_events(run_id),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


def _safe_name(value: str) -> str:
    text = str(value or "").strip()
    if text == "":
        return "NA"
    return "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in text)


def _parse_iso_datetime(value: str) -> datetime:
    text = str(value or "").strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    dt = datetime.fromisoformat(text)
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _norm_freq(value: object) -> str:
    try:
        return f"{float(value):.6f}"
    except Exception:
        return str(value)


@router.post("/calibration/reports/generate", response_model=CalibrationReportGenerateResponse)
def generate_calibration_report(payload: CalibrationReportGenerateRequest):
    normalized_cal_type = str(payload.cal_type or "").strip().lower()
    repo_by_type = {
        "cal_sg": (CalSgCalibrationRepository, settings.CAL_SG_CALIBRATION_TABLE),
        "inject_cal": (InjectCalCalibrationRepository, settings.INJECT_CAL_CALIBRATION_TABLE),
        "downlink": (DownlinkCalCalibrationRepository, settings.DOWNLINK_CAL_CALIBRATION_TABLE),
    }
    repo_config = repo_by_type.get(normalized_cal_type)
    if repo_config is None:
        raise HTTPException(status_code=400, detail="Report generation currently supports cal_sg, inject_cal, and downlink")

    cal_id = str(payload.cal_id or "").strip()
    if cal_id == "":
        raise HTTPException(status_code=400, detail=f"cal_id is required for {normalized_cal_type} report generation")

    repo_class, table_name = repo_config
    cal_repo = repo_class(Database._db_path, table_name)
    env_repo = EnvDataRepository(Database._db_path, "EnvData")

    rows = cal_repo.list_rows(cal_id=cal_id)
    if len(rows) == 0:
        raise HTTPException(status_code=404, detail=f"No {normalized_cal_type} calibration data found for cal_id: {cal_id}")

    env_values = {row["parameter"]: row["value"] for row in env_repo.list_rows()}
    satellite_name = str(env_values.get("SATELLITE_NAME", "")).strip()
    results_directory = str(env_values.get("RESULTS_DIRECTORY", "")).strip()

    if results_directory == "":
        raise HTTPException(status_code=400, detail="RESULTS_DIRECTORY is empty in ENV Data")

    base_dir = Path(results_directory).expanduser()
    calibration_data_dir = base_dir / "CalibrationData"
    pdf_dir = calibration_data_dir / "pdf"
    excels_dir = calibration_data_dir / "excels"
    pdf_dir.mkdir(parents=True, exist_ok=True)
    excels_dir.mkdir(parents=True, exist_ok=True)

    latest_dt = max(_parse_iso_datetime(str(row["datetime"])) for row in rows)
    stamp = latest_dt.strftime("%Y%m%d_%H%M%S")
    sat_safe = _safe_name(satellite_name) if satellite_name else "unknown"
    cal_type_safe = _safe_name(normalized_cal_type)
    base_name = f"{sat_safe}_{cal_type_safe}_{stamp}"

    pdf_path = pdf_dir / f"{base_name}.pdf"
    excel_path = excels_dir / f"{base_name}.xlsx"

    pdf_generated = False
    if not pdf_path.exists():
        try:
            from jinja2 import Environment, FileSystemLoader, select_autoescape
        except Exception as exc:
            raise HTTPException(
                status_code=500,
                detail=f"PDF dependencies missing. Install jinja2. Error: {exc}",
            ) from exc

        # Use bundled WeasyPrint wrapper executable from known local paths.
        _project_root = Path(__file__).parent.parent.parent
        wrapper_candidates = [
            Path(r"E:\Code\Mainframe\DotNet\PdfApp\bin\Release\net10.0\win-x64\publish\PdfApp.exe"),
            _project_root / "weasyprintWrapper" / "net10.0" / "PdfApp.exe",
            _project_root / "weasyprintWrapper" / "net10.0" / "win-x64" / "publish" / "PdfApp.exe",
            _project_root / "weasyprintWrapper" / "net10.0" / "win-x64" / "PdfApp.exe",
        ]
        wrapper_exe = next((p for p in wrapper_candidates if p.is_file()), None)
        if wrapper_exe is None:
            raise HTTPException(
                status_code=500,
                detail=(
                    "WeasyPrint wrapper executable not found. Checked: "
                    + "; ".join(str(p) for p in wrapper_candidates)
                ),
            )

        templates_dir = Path(__file__).parent.parent / "templates"
        env = Environment(
            loader=FileSystemLoader(str(templates_dir)),
            autoescape=select_autoescape(["html"]),
        )

        if normalized_cal_type == "inject_cal":
            template_name = "inject_cal_report.html"
            title = "Inject Calibration"
            template_rows = [
                {
                    "frequency": float(row["frequency"]),
                    "sa_loss": float(row["sa_loss"]),
                    "dl_pm_loss": float(row["dl_pm_loss"]),
                    "datetime": str(row["datetime"]),
                }
                for row in rows
            ]
            template_spur_rows = []
        elif normalized_cal_type == "downlink":
            template_name = "downlink_cal_report.html"
            title = "Downlink Calibration"
            template_rows = []
            template_spur_rows = []
            for row in rows:
                mapped = {
                    "code": str(row.get("code", "")),
                    "port": str(row.get("port", "")),
                    "frequency": float(row["frequency"]),
                    "frequency_label": str(row.get("frequency_label", "")),
                    "value": float(row["value"]),
                    "datetime": str(row["datetime"]),
                }
                if str(mapped["frequency_label"]).strip().lower() == "spur_band":
                    template_spur_rows.append(mapped)
                else:
                    template_rows.append(mapped)
        else:
            template_name = "cal_sg_report.html"
            title = "Link Loss Calibration"
            template_rows = [
                {
                    "frequency": float(row["frequency"]),
                    "value": float(row["value"]),
                    "datetime": str(row["datetime"]),
                }
                for row in rows
            ]
            template_spur_rows = []

        template = env.get_template(template_name)
        html_content = template.render(
            title=title,
            cal_type=normalized_cal_type,
            satellite_name=satellite_name or "N/A",
            generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            data_timestamp=latest_dt.strftime("%Y-%m-%d %H:%M:%S UTC"),
            rows=template_rows,
            spur_rows=template_spur_rows,
        )

        temp_html_path: Path | None = None
        try:
            with tempfile.NamedTemporaryFile(
                mode="w",
                suffix=".html",
                prefix=f"{base_name}_",
                dir=str(pdf_dir),
                encoding="utf-8",
                delete=False,
            ) as temp_html:
                temp_html.write(html_content)
                temp_html_path = Path(temp_html.name)

            proc = subprocess.run(
                [str(wrapper_exe), str(temp_html_path), str(pdf_path)],
                capture_output=True,
                text=True,
                check=False,
                cwd=str(wrapper_exe.parent),
            )
            stdout_text = (proc.stdout or "").strip()
            stderr_text = (proc.stderr or "").strip()
            if proc.returncode != 0:
                raise HTTPException(
                    status_code=500,
                    detail=(
                        "PDF generation failed via WeasyPrint wrapper. "
                        f"exit={proc.returncode}; stdout={stdout_text}; stderr={stderr_text}"
                    ),
                )

            if "weasyprint error" in stderr_text.lower() or "weasyprint error" in stdout_text.lower():
                raise HTTPException(
                    status_code=500,
                    detail=(
                        "WeasyPrint wrapper reported an error. "
                        f"stdout={stdout_text}; stderr={stderr_text}"
                    ),
                )

            if not pdf_path.exists() or pdf_path.stat().st_size == 0:
                generated_path: Path | None = None
                for line in (stdout_text + "\n" + stderr_text).splitlines():
                    if line.lower().startswith("pdf generated at:"):
                        candidate = line.split(":", 1)[1].strip()
                        p = Path(candidate)
                        if p.exists() and p.is_file():
                            generated_path = p
                            break

                if generated_path is not None and generated_path.resolve() != pdf_path.resolve():
                    shutil.copyfile(str(generated_path), str(pdf_path))

                if not pdf_path.exists() or pdf_path.stat().st_size == 0:
                    raise HTTPException(
                        status_code=500,
                        detail=(
                            "WeasyPrint wrapper completed but no PDF was produced at expected path. "
                            f"expected={pdf_path}; stdout={stdout_text}; stderr={stderr_text}"
                        ),
                    )
        except Exception as exc:
            if isinstance(exc, HTTPException):
                raise
            raise HTTPException(status_code=500, detail=f"PDF generation failed. Error: {exc}") from exc
        finally:
            if temp_html_path and temp_html_path.exists():
                try:
                    temp_html_path.unlink()
                except Exception:
                    pass
        pdf_generated = True

    try:
        from openpyxl import Workbook, load_workbook
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Excel dependencies missing. Install openpyxl. Error: {exc}",
        ) from exc

    if excel_path.exists():
        workbook = load_workbook(str(excel_path))
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)

    appended = 0

    if normalized_cal_type == "inject_cal":
        sheet = workbook["Calibration"] if "Calibration" in workbook.sheetnames else workbook.create_sheet("Calibration")
        if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
            sheet.append(["caltype", "freq(Mhz)", "SA Loss(dB)", "DL_PM Loss(dB)", "Datetime"])

        existing_keys = set()
        for row_idx in range(2, sheet.max_row + 1):
            freq = sheet.cell(row=row_idx, column=2).value
            dt = sheet.cell(row=row_idx, column=5).value
            existing_keys.add((_norm_freq(freq), str(dt)))

        for row in rows:
            freq_text = _norm_freq(row["frequency"])
            dt_text = str(row["datetime"])
            key = (freq_text, dt_text)
            if key in existing_keys:
                continue
            sheet.append([
                normalized_cal_type,
                float(row["frequency"]),
                float(row["sa_loss"]),
                float(row["dl_pm_loss"]),
                dt_text,
            ])
            existing_keys.add(key)
            appended += 1
    elif normalized_cal_type == "downlink":
        normal_sheet = workbook["Downlink"] if "Downlink" in workbook.sheetnames else workbook.create_sheet("Downlink")
        spur_sheet = workbook["Downlink_SpurBand"] if "Downlink_SpurBand" in workbook.sheetnames else workbook.create_sheet("Downlink_SpurBand")

        if normal_sheet.max_row == 1 and normal_sheet.cell(row=1, column=1).value is None:
            normal_sheet.append(["caltype", "code", "port", "freq(Mhz)", "frequency_label", "value(dB)", "Datetime"])
        if spur_sheet.max_row == 1 and spur_sheet.cell(row=1, column=1).value is None:
            spur_sheet.append(["caltype", "code", "port", "freq(Mhz)", "frequency_label", "value(dB)", "Datetime"])

        normal_keys = set()
        for row_idx in range(2, normal_sheet.max_row + 1):
            normal_keys.add((
                str(normal_sheet.cell(row=row_idx, column=2).value or "").strip(),
                str(normal_sheet.cell(row=row_idx, column=3).value or "").strip(),
                _norm_freq(normal_sheet.cell(row=row_idx, column=4).value),
                str(normal_sheet.cell(row=row_idx, column=5).value or "").strip().lower(),
                str(normal_sheet.cell(row=row_idx, column=7).value),
            ))

        spur_keys = set()
        for row_idx in range(2, spur_sheet.max_row + 1):
            spur_keys.add((
                str(spur_sheet.cell(row=row_idx, column=2).value or "").strip(),
                str(spur_sheet.cell(row=row_idx, column=3).value or "").strip(),
                _norm_freq(spur_sheet.cell(row=row_idx, column=4).value),
                str(spur_sheet.cell(row=row_idx, column=5).value or "").strip().lower(),
                str(spur_sheet.cell(row=row_idx, column=7).value),
            ))

        for row in rows:
            code = str(row.get("code", "")).strip()
            port = str(row.get("port", "")).strip()
            frequency = float(row["frequency"])
            frequency_label = str(row.get("frequency_label", "")).strip()
            dt_text = str(row["datetime"])
            key = (code, port, _norm_freq(frequency), frequency_label.lower(), dt_text)
            target_is_spur = frequency_label.lower() == "spur_band"
            if target_is_spur:
                if key in spur_keys:
                    continue
                spur_sheet.append([
                    normalized_cal_type,
                    code,
                    port,
                    frequency,
                    frequency_label,
                    float(row["value"]),
                    dt_text,
                ])
                spur_keys.add(key)
            else:
                if key in normal_keys:
                    continue
                normal_sheet.append([
                    normalized_cal_type,
                    code,
                    port,
                    frequency,
                    frequency_label,
                    float(row["value"]),
                    dt_text,
                ])
                normal_keys.add(key)
            appended += 1
    else:
        sheet = workbook["Calibration"] if "Calibration" in workbook.sheetnames else workbook.create_sheet("Calibration")
        if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
            sheet.append(["caltype", "freq(Mhz)", "Loss(dB)", "Datetime"])

        existing_keys = set()
        for row_idx in range(2, sheet.max_row + 1):
            freq = sheet.cell(row=row_idx, column=2).value
            dt = sheet.cell(row=row_idx, column=4).value
            existing_keys.add((_norm_freq(freq), str(dt)))

        for row in rows:
            freq_text = _norm_freq(row["frequency"])
            dt_text = str(row["datetime"])
            key = (freq_text, dt_text)
            if key in existing_keys:
                continue
            sheet.append([
                normalized_cal_type,
                float(row["frequency"]),
                float(row["value"]),
                dt_text,
            ])
            existing_keys.add(key)
            appended += 1

    workbook.save(str(excel_path))

    message = "Report generated"
    if not pdf_generated and appended == 0:
        message = "No data change detected; existing report files kept"
    elif not pdf_generated:
        message = "PDF already exists for this dataset; Excel updated"

    return CalibrationReportGenerateResponse(
        cal_id=cal_id,
        cal_type=normalized_cal_type,
        satellite_name=satellite_name,
        results_directory=str(base_dir),
        calibration_data_directory=str(calibration_data_dir),
        pdf_path=str(pdf_path),
        excel_path=str(excel_path),
        pdf_generated=pdf_generated,
        excel_rows_appended=appended,
        message=message,
    )
